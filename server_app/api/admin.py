from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import path
from django.utils.html import format_html
from django import forms
from django.views.decorators.http import require_http_methods
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
from .models import WorkLog, Task, EmployeeProfile, GlobalSettings, TaskAttachment

# --- 1. ВСТРАИВАЕМ ПРОФИЛЬ И ИСПРАВЛЯЕМ ОТОБРАЖЕНИЕ ЮЗЕРА ---
class EmployeeInline(admin.StackedInline):
    model = EmployeeProfile
    can_delete = False
    verbose_name_plural = 'Данные сотрудника (Табельный номер, Должность)'

class UserAdmin(BaseUserAdmin):
    inlines = (EmployeeInline,)
    list_display = ('username', 'get_full_name_custom', 'get_personnel_number', 'is_active')
    
    def get_full_name_custom(self, obj):
        return obj.get_full_name()
    get_full_name_custom.short_description = "ФИО (Русский)"

    def get_personnel_number(self, obj):
        if hasattr(obj, 'profile'):
            return obj.profile.personnel_number
        return "-"
    get_personnel_number.short_description = "Табельный №"

admin.site.unregister(User)
admin.site.register(User, UserAdmin)

# --- 2. ФОРМА ДЛЯ ОТЧЕТА ---
class ReportForm(forms.Form):
    EMPLOYEE_CHOICES = [('all', 'Все сотрудники')]
    
    start_date = forms.DateField(label="Начальная дата", widget=forms.DateInput(attrs={'type': 'date'}))
    end_date = forms.DateField(label="Конечная дата", widget=forms.DateInput(attrs={'type': 'date'}))
    responsible_person = forms.CharField(label="Ответственный за отчет (ФИО)", max_length=100)
    company_name = forms.CharField(label="Название компании", max_length=200, required=False)
    employees = forms.MultipleChoiceField(
        label="Сотрудники",
        choices=EMPLOYEE_CHOICES,
        widget=forms.CheckboxSelectMultiple(),
        required=False,
        help_text="Выберите сотрудников для включения в отчет. Если не выбрано, будут включены все."
    )
    include_completed_tasks = forms.BooleanField(
        label="Добавить колонку 'Выполнено задач за период'",
        required=False,
        initial=False
    )
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Динамически добавляем сотрудников в choices
        from .models import EmployeeProfile
        employee_choices = [('all', 'Все сотрудники')]
        for emp in EmployeeProfile.objects.filter(user__is_staff=False).select_related('user'):
            employee_choices.append((str(emp.id), f"{emp.personnel_number} - {emp.user.get_full_name() or emp.user.username}"))
        self.fields['employees'].choices = employee_choices

# --- 2.1 ФОРМА ДЛЯ РЕДАКТИРОВАНИЯ ПРОФИЛЯ СОТРУДНИКА ---
class EmployeeProfileEditForm(forms.ModelForm):
    first_name = forms.CharField(label="Имя", max_length=100, required=False, 
                                 widget=forms.TextInput(attrs={'class': 'form-control'}))
    last_name = forms.CharField(label="Фамилия", max_length=100, required=False,
                               widget=forms.TextInput(attrs={'class': 'form-control'}))
    email = forms.EmailField(label="Email", required=False,
                            widget=forms.EmailInput(attrs={'class': 'form-control'}))
    
    class Meta:
        model = EmployeeProfile
        fields = [
            'position',
            'department',
            'anti_afk_enabled',
            'anti_afk_idle_minutes',
            'anti_afk_grace_seconds'
        ]
        labels = {
            'position': 'Должность',
            'department': 'Отдел',
            'anti_afk_enabled': 'Включить Anti-AFK',
            'anti_afk_idle_minutes': 'Минут бездействия до предупреждения',
            'anti_afk_grace_seconds': 'Секунды обратного отсчета',
        }
        widgets = {
            'position': forms.TextInput(attrs={'class': 'form-control'}),
            'department': forms.TextInput(attrs={'class': 'form-control'}),
            'anti_afk_enabled': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'anti_afk_idle_minutes': forms.NumberInput(attrs={'class': 'form-control', 'min': '1'}),
            'anti_afk_grace_seconds': forms.NumberInput(attrs={'class': 'form-control', 'min': '1'}),
        }

# --- 3. АДМИНКА ДЛЯ ЛОГОВ С ПРОМЕЖУТОЧНОЙ СТРАНИЦЕЙ ---
@admin.register(WorkLog)
class WorkLogAdmin(admin.ModelAdmin):
    list_display = ('user', 'event_colored', 'time_fmt', 'created_at')
    list_filter = ('user', 'event')
    change_list_template = "admin/worklog_change_list.html" # Используем стандартный шаблон, но добавим кнопку через actions

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('export-report/', self.admin_site.admin_view(self.export_report_view), name='export_report'),
        ]
        return my_urls + urls

    def export_report_view(self, request):
        # Получаем настройки компании или дефолт
        settings = GlobalSettings.objects.first()
        default_company = settings.company_name if settings else "ООО Моя Компания"
        
        if request.method == 'POST':
            form = ReportForm(request.POST)
            if form.is_valid():
                return self.generate_excel(form.cleaned_data)
        else:
            # Предзаполняем форму: по умолчанию ответственный - текущий админ
            initial_data = {
                'start_date': datetime.now().date() - timedelta(days=7),
                'end_date': datetime.now().date(),
                'responsible_person': request.user.get_full_name() or request.user.username,
                'company_name': default_company,
                'include_completed_tasks': False
            }
            form = ReportForm(initial=initial_data)

        context = dict(
           self.admin_site.each_context(request),
           form=form,
           title="Выгрузка табеля учета времени"
        )
        return render(request, "admin/export_form.html", context)

    def generate_excel(self, data):
        # Сохраняем название компании на будущее
        settings, created = GlobalSettings.objects.get_or_create(id=1)
        settings.company_name = data['company_name']
        settings.save()

        # Фильтруем логи по фактическому времени события (timestamp),
        # а не по created_at (время вставки записи в БД).
        start_dt = datetime.combine(data['start_date'], datetime.min.time())
        end_dt = datetime.combine(data['end_date'], datetime.max.time())
        start_ts = start_dt.timestamp()
        end_ts = end_dt.timestamp()

        logs = WorkLog.objects.filter(
            timestamp__range=(start_ts, end_ts)
        ).order_by('user', 'timestamp')

        include_completed_tasks = data.get('include_completed_tasks', False)
        end_col_letter = 'G' if include_completed_tasks else 'F'
        date_col_letter = 'G' if include_completed_tasks else 'F'

        # Считаем выполненные задачи за период по сотрудникам.
        completed_tasks_map = {}
        if include_completed_tasks:
            for row in Task.objects.filter(
                is_completed=True,
                created_at__range=(start_dt, end_dt)
            ).values('user_id'):
                uid = row['user_id']
                completed_tasks_map[uid] = completed_tasks_map.get(uid, 0) + 1

        # === ГЕНЕРАЦИЯ EXCEL (Тот же код, но с переменными из формы) ===
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Табель"
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(f'A1:{end_col_letter}1')
        ws['A1'] = f"ТАБЕЛЬ УЧЕТА ({data['company_name']})"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center_align

        ws.merge_cells(f'A2:{end_col_letter}2')
        ws['A2'] = f"Период: с {data['start_date'].strftime('%d.%m.%Y')} по {data['end_date'].strftime('%d.%m.%Y')}"
        ws['A2'].alignment = center_align

        ws['A3'] = f"Ответственный: {data['responsible_person']}"
        ws[f'{date_col_letter}3'] = datetime.now().strftime('%d.%m.%Y')

        headers = ['Таб. №', 'ФИО Сотрудника', 'Должность', 'Дата', 'Интервал', 'Часов']
        if include_completed_tasks:
            headers.append('Выполнено задач')
        ws.append([])
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = bold_font
            cell.border = thin_border

        # -- ЛОГИКА ПОДСЧЕТА --
        data_map = {}
        for log in logs:
            uid = log.user.id
            d_str = datetime.fromtimestamp(log.timestamp).strftime('%Y-%m-%d')
            if uid not in data_map: data_map[uid] = {}
            if d_str not in data_map[uid]: data_map[uid][d_str] = []
            data_map[uid][d_str].append(log)

        current_row = 6
        for uid, dates in data_map.items():
            user = User.objects.get(id=uid)
            try:
                prof = user.profile
                fio = user.get_full_name() or user.username
                tnum = prof.personnel_number
                pos = prof.position
            except:
                fio = user.username
                tnum = "-"
                pos = "-"
            tasks_done = completed_tasks_map.get(uid, 0)

            total_hours = 0
            for date_str, events in dates.items():
                daily_sec = 0
                temp_start = None
                s_str = datetime.fromtimestamp(events[0].timestamp).strftime('%H:%M')
                e_str = datetime.fromtimestamp(events[-1].timestamp).strftime('%H:%M')

                for ev in events:
                    if ev.event.upper() in ['START', 'RESUME']:
                        temp_start = ev.timestamp
                    elif ev.event.upper() in ['STOP', 'PAUSE'] and temp_start:
                        daily_sec += (ev.timestamp - temp_start)
                        temp_start = None
                
                hours = round(daily_sec / 3600, 2)
                total_hours += hours
                
                row_data = [tnum, fio, pos, date_str, f"{s_str}-{e_str}", hours]
                if include_completed_tasks:
                    row_data.append(tasks_done)
                ws.append(row_data)
                current_row += 1

            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.cell(row=current_row, column=1).value = f"ИТОГО {fio}:"
            ws.cell(row=current_row, column=6).value = total_hours
            ws.cell(row=current_row, column=6).font = bold_font
            if include_completed_tasks:
                ws.cell(row=current_row, column=7).value = tasks_done
                ws.cell(row=current_row, column=7).font = bold_font
            current_row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Timesheet_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        wb.save(response)
        return response

    # Добавляем кнопку в UI админки
    actions = ['go_to_export']
    def go_to_export(self, request, queryset):
        return HttpResponseRedirect("export-report/")
    go_to_export.short_description = "📊 Перейти к выгрузке Табеля"

    def event_colored(self, obj):
        colors = {"start": "green", "stop": "red", "pause": "orange", "resume": "blue", "START": "green", "STOP": "red", "PAUSE": "orange", "RESUME": "blue"}
        return format_html('<span style="color: {};"><b>{}</b></span>', colors.get(obj.event, "black"), obj.event)

    def time_fmt(self, obj):
        return datetime.fromtimestamp(obj.timestamp).strftime('%d.%m.%Y %H:%M:%S')

# --- INLINE ДЛЯ ВЛОЖЕНИЙ К ЗАДАЧАМ ---
class TaskAttachmentInline(admin.TabularInline):
    model = TaskAttachment
    extra = 0
    fields = ('original_filename', 'file_size_mb', 'uploaded_by', 'created_at', 'file_actions')
    readonly_fields = ('original_filename', 'file_size_mb', 'uploaded_by', 'created_at', 'file_actions')
    can_delete = True
    
    def file_size_mb(self, obj):
        """Показать размер в МБ"""
        if obj.file_size:
            return f"{round(obj.file_size / (1024 * 1024), 2)} МБ"
        return "-"
    file_size_mb.short_description = "Размер"
    
    def file_actions(self, obj):
        """Ссылки для скачивания"""
        if not obj or not obj.pk:
            return "—"
        download_url = f"/admin/api/task/attachments/{obj.pk}/view/"
        return format_html(
            '<a class="button" href="{}">⬇ Скачать</a>',
            download_url,
        )
    file_actions.short_description = "Действия"
    
    def has_add_permission(self, request, obj=None):
        # Вложения могут добавляться только через API, не через админку
        return False

@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    list_display = ('title', 'user', 'deadline', 'is_completed', 'has_attachments')
    list_filter = ('is_completed', 'user')
    readonly_fields = ('created_at', 'created_by')
    fieldsets = (
        ('Основная информация', {
            'fields': ('title', 'description', 'user', 'deadline', 'is_completed')
        }),
        ('Служебная информация', {
            'fields': ('created_at', 'created_by')
        }),
    )
    inlines = [TaskAttachmentInline]
    
    def has_attachments(self, obj):
        """Показать иконку если есть вложения"""
        count = obj.attachments.count()
        if count > 0:
            return format_html('📎 {}', count)
        return '-'
    has_attachments.short_description = "Файлы"
    
    def get_urls(self):
        """Добавить кастомные URLs для работы с вложениями"""
        urls = super().get_urls()
        custom_urls = [
            path('attachments/<int:attachment_id>/view/', 
                 self.admin_site.admin_view(self.view_attachment_view), 
                 name='view_attachment'),
            path('attachments/<int:attachment_id>/delete/', 
                 self.admin_site.admin_view(self.delete_attachment_view), 
                 name='delete_attachment'),
        ]
        return custom_urls + urls
    
    def view_attachment_view(self, request, attachment_id):
        """Скачать вложение"""
        from django.http import FileResponse
        attachment = get_object_or_404(TaskAttachment, id=attachment_id)
        
        if not attachment.file:
            return HttpResponse('Файл не найден', status=404)
        
        return FileResponse(
            attachment.file.open("rb"),
            as_attachment=True,
            filename=attachment.original_filename,
        )
    
    def delete_attachment_view(self, request, attachment_id):
        """Удалить вложение"""
        attachment = get_object_or_404(TaskAttachment, id=attachment_id)
        task_id = attachment.task.id
        
        if attachment.file:
            attachment.file.delete()
        attachment.delete()
        
        # Вернуться обратно на страницу задачи
        from django.contrib.admin.helpers import AdminReadonlyField
        return redirect(f'/admin/api/task/{task_id}/change/')

@admin.register(EmployeeProfile)
class EmployeeProfileAdmin(admin.ModelAdmin):
    list_display = (
        'personnel_number',
        'get_full_name',
        'position',
        'department',
        'get_email',
        'get_status',
        'anti_afk_enabled',
        'anti_afk_idle_minutes',
        'anti_afk_grace_seconds',
    )
    list_filter = ('department', 'anti_afk_enabled')
    search_fields = ('personnel_number', 'user__first_name', 'user__last_name', 'user__email')
    readonly_fields = ('last_activity',)
    change_list_template = "admin/employees_change_list.html"
    
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('export-report/', self.admin_site.admin_view(self.export_employee_report_view), name='export_employee_report'),
            path('<int:employee_id>/edit/', self.admin_site.admin_view(self.edit_employee_view), name='edit_employee'),
            path('<int:employee_id>/export-report/', self.admin_site.admin_view(self.export_single_employee_report_view), name='export_single_employee_report'),
        ]
        return my_urls + urls
    
    def get_full_name(self, obj):
        return obj.user.get_full_name() or obj.user.username
    get_full_name.short_description = "ФИО"
    
    def get_email(self, obj):
        return obj.user.email or "-"
    get_email.short_description = "Email"
    
    def get_status(self, obj):
        status = "🟢 Онлайн" if obj.is_online() else "⚫ Оффлайн"
        return format_html('<span style="font-size: 14px;">{}</span>', status)
    get_status.short_description = "Статус"
    
    def get_queryset(self, request):
        # Исключаем администраторов
        qs = super().get_queryset(request)
        return qs.filter(user__is_staff=False)
    
    def edit_employee_view(self, request, employee_id):
        """Страница редактирования профиля сотрудника"""
        employee = get_object_or_404(EmployeeProfile, id=employee_id)
        
        if request.method == 'POST':
            form = EmployeeProfileEditForm(request.POST, instance=employee)
            if form.is_valid():
                # Сохраняем данные профиля
                employee_obj = form.save()
                
                # Обновляем данные User если были заполнены
                first_name = form.cleaned_data.get('first_name', '')
                last_name = form.cleaned_data.get('last_name', '')
                email = form.cleaned_data.get('email', '')
                
                user = employee_obj.user
                if first_name:
                    user.first_name = first_name
                if last_name:
                    user.last_name = last_name
                if email:
                    user.email = email
                user.save()
                
                # Перенаправляем обратно в список
                return redirect('admin:api_employeeprofile_changelist')
        else:
            initial_data = {
                'first_name': employee.user.first_name,
                'last_name': employee.user.last_name,
                'email': employee.user.email,
            }
            form = EmployeeProfileEditForm(instance=employee, initial=initial_data)
        
        context = dict(
            self.admin_site.each_context(request),
            employee=employee,
            form=form,
            title=f"Редактирование профиля {employee.user.get_full_name() or employee.user.username}",
            opts=self.model._meta,
        )
        return render(request, "admin/employee_edit.html", context)
    
    def export_single_employee_report_view(self, request, employee_id):
        """Выгрузка отчёта только для одного сотрудника"""
        employee = get_object_or_404(EmployeeProfile, id=employee_id)
        user = employee.user
        
        # Получаем параметры из GET запроса или используем дефолты
        today = datetime.now().date()
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Парсим даты если они переданы
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except:
                start_date = today - timedelta(days=7)
        else:
            start_date = today - timedelta(days=7)
            
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except:
                end_date = today
        else:
            end_date = today
        
        # Фильтруем логи только для этого пользователя
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        start_ts = start_dt.timestamp()
        end_ts = end_dt.timestamp()

        logs = WorkLog.objects.filter(
            user=user,
            timestamp__range=(start_ts, end_ts)
        ).order_by('timestamp')
        
        # Получаем название компании
        settings = GlobalSettings.objects.first()
        company_name = settings.company_name if settings else "ООО Моя Компания"
        
        # === ГЕНЕРАЦИЯ EXCEL ===
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Табель"
        
        bold_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='417690', end_color='417690', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = f"ТАБЕЛЬ УЧЕТА ВРЕМЕНИ - {employee.user.get_full_name() or employee.user.username}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center_align
        
        # Компания
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Компания: {company_name}"
        ws['A2'].alignment = center_align
        
        # Период
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Период: с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
        ws['A3'].alignment = center_align
        
        # Информация сотрудника
        ws['A4'] = f"Табельный номер:"
        ws['B4'] = employee.personnel_number
        ws['D4'] = f"Должность:"
        ws['E4'] = employee.position
        
        ws['A5'] = f"Отдел:"
        ws['B5'] = employee.department or "-"
        ws['D5'] = f"Email:"
        ws['E5'] = user.email or "-"
        
        # Заголовки таблицы
        header_row = 7
        headers = ['Дата', 'Начало', 'Окончание', 'Часов', 'Событие', 'Примечание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Заполняем данные
        current_row = header_row + 1
        data_map = {}
        
        for log in logs:
            d_str = datetime.fromtimestamp(log.timestamp).strftime('%Y-%m-%d')
            if d_str not in data_map:
                data_map[d_str] = {'events': [], 'start': None, 'end': None, 'daily_sec': 0}
            
            data_map[d_str]['events'].append(log)
            
            if log.event.upper() in ['START', 'RESUME']:
                data_map[d_str]['start'] = datetime.fromtimestamp(log.timestamp)
            elif log.event.upper() in ['STOP', 'PAUSE']:
                data_map[d_str]['end'] = datetime.fromtimestamp(log.timestamp)
        
        # Обработка логов по датам
        total_hours = 0
        for date_str in sorted(data_map.keys()):
            data = data_map[date_str]
            events = data['events']
            
            daily_sec = 0
            temp_start = None
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            
            for ev in events:
                if ev.event.upper() in ['START', 'RESUME']:
                    temp_start = ev.timestamp
                elif ev.event.upper() in ['STOP', 'PAUSE'] and temp_start:
                    daily_sec += (ev.timestamp - temp_start)
                    temp_start = None
            
            hours = round(daily_sec / 3600, 2)
            total_hours += hours
            
            # Определяем начало и конец дня
            start_time = events[0].timestamp
            end_time = events[-1].timestamp
            start_str = datetime.fromtimestamp(start_time).strftime('%H:%M')
            end_str = datetime.fromtimestamp(end_time).strftime('%H:%M')
            
            # Записываем в точки
            ws.cell(row=current_row, column=1).value = date_obj.strftime('%d.%m.%Y')
            ws.cell(row=current_row, column=2).value = start_str
            ws.cell(row=current_row, column=3).value = end_str
            ws.cell(row=current_row, column=4).value = hours
            ws.cell(row=current_row, column=5).value = "Работа"
            ws.cell(row=current_row, column=6).value = employee.personnel_number
            
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = thin_border
            
            current_row += 1
        
        # Итого
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1).value = "ИТОГО ЧАСОВ:"
        ws.cell(row=current_row, column=1).font = bold_font
        ws.cell(row=current_row, column=5).value = total_hours
        ws.cell(row=current_row, column=5).font = bold_font
        
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border
        
        # Установка ширины столбцов
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        # Возвращаем файл
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Timesheet_{employee.personnel_number}_{start_date}_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def export_employee_report_view(self, request):
        settings = GlobalSettings.objects.first()
        default_company = settings.company_name if settings else "ООО Моя Компания"
        
        if request.method == 'POST':
            form = ReportForm(request.POST)
            if form.is_valid():
                return self.generate_employee_excel(form.cleaned_data)
        else:
            initial_data = {
                'start_date': datetime.now().date() - timedelta(days=7),
                'end_date': datetime.now().date(),
                'responsible_person': request.user.get_full_name() or request.user.username,
                'company_name': default_company,
                'include_completed_tasks': False
            }
            form = ReportForm(initial=initial_data)

        context = dict(
           self.admin_site.each_context(request),
           form=form,
           title="Выгрузка табеля учета времени"
        )
        return render(request, "admin/export_form.html", context)
    
    def generate_employee_excel(self, data):
        """Генерирует профессиональный отчёт для бухгалтерии с выбранными сотрудниками"""
        settings, created = GlobalSettings.objects.get_or_create(id=1)
        settings.company_name = data['company_name']
        settings.save()

        start_dt = datetime.combine(data['start_date'], datetime.min.time())
        end_dt = datetime.combine(data['end_date'], datetime.max.time())
        
        # Определяем фильтр по сотрудникам
        employees_ids = data.get('employees', [])
        if not employees_ids or 'all' in employees_ids:
            # Получаем всех сотрудников
            emp_list = EmployeeProfile.objects.filter(user__is_staff=False).select_related('user')
        else:
            # Получаем выбранных сотрудников
            emp_list = EmployeeProfile.objects.filter(id__in=employees_ids, user__is_staff=False).select_related('user')
        
        start_ts = start_dt.timestamp()
        end_ts = end_dt.timestamp()
        include_completed_tasks = data.get('include_completed_tasks', False)
        end_col_letter = 'G' if include_completed_tasks else 'F'
        date_col_letter = 'G' if include_completed_tasks else 'F'

        logs = WorkLog.objects.filter(
            timestamp__range=(start_ts, end_ts),
            user__in=[e.user_id for e in emp_list]
        ).order_by('user', 'timestamp')

        completed_tasks_map = {}
        if include_completed_tasks:
            for row in Task.objects.filter(
                is_completed=True,
                created_at__range=(start_dt, end_dt),
                user__in=[e.user_id for e in emp_list]
            ).values('user_id'):
                uid = row['user_id']
                completed_tasks_map[uid] = completed_tasks_map.get(uid, 0) + 1

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Табель"
        
        # Стили
        title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        bold_font = Font(name='Calibri', size=11, bold=True)
        regular_font = Font(name='Calibri', size=11)
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Заголовок (тёмно-синий)
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        
        # Серый фон для заголовков таблицы
        table_header_fill = PatternFill(start_color='417690', end_color='417690', fill_type='solid')
        
        # Светлый фон для итогов
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        if include_completed_tasks:
            ws.column_dimensions['G'].width = 18

        # Заголовок документа
        ws.merge_cells(f'A1:{end_col_letter}1')
        cell = ws['A1']
        cell.value = f"ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ"
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.row_dimensions[1].height = 25

        # Информация о компании и периоде
        ws.merge_cells(f'A2:{end_col_letter}2')
        cell = ws['A2']
        cell.value = f"{data['company_name']}"
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.alignment = center_align

        ws.merge_cells(f'A3:{end_col_letter}3')
        cell = ws['A3']
        cell.value = f"Период: с {data['start_date'].strftime('%d.%m.%Y')} по {data['end_date'].strftime('%d.%m.%Y')}"
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = center_align

        # Подписи (строка 4)
        ws['A4'] = f"Ответственный: {data['responsible_person']}"
        ws['A4'].font = regular_font
        ws[f'{date_col_letter}4'] = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"
        ws[f'{date_col_letter}4'].font = regular_font
        ws[f'{date_col_letter}4'].alignment = right_align

        # Пустая строка
        ws.append([])
        
        # Заголовки таблицы (строка 6)
        headers = ['Таб. №', 'ФИО Сотрудника', 'Должность', 'Дата', 'Время (начало-конец)', 'Часов']
        if include_completed_tasks:
            headers.append('Выполнено задач')
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = table_header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        ws.row_dimensions[6].height = 20

        # Подготовка данных
        data_map = {}
        for log in logs:
            uid = log.user.id
            d_str = datetime.fromtimestamp(log.timestamp).strftime('%Y-%m-%d')
            if uid not in data_map:
                data_map[uid] = {}
            if d_str not in data_map[uid]:
                data_map[uid][d_str] = []
            data_map[uid][d_str].append(log)

        # Заполнение данных
        current_row = 7
        for uid, dates in sorted(data_map.items()):
            user = User.objects.get(id=uid)
            try:
                prof = user.profile
                fio = user.get_full_name() or user.username
                tnum = prof.personnel_number
                pos = prof.position
            except:
                fio = user.username
                tnum = "-"
                pos = "-"
            tasks_done = completed_tasks_map.get(uid, 0)

            total_hours = 0
            dates_list = sorted(dates.items())
            
            for date_idx, (date_str, events) in enumerate(dates_list):
                daily_sec = 0
                temp_start = None
                s_str = datetime.fromtimestamp(events[0].timestamp).strftime('%H:%M')
                e_str = datetime.fromtimestamp(events[-1].timestamp).strftime('%H:%M')

                for ev in events:
                    if ev.event.upper() in ['START', 'RESUME']:
                        temp_start = ev.timestamp
                    elif ev.event.upper() in ['STOP', 'PAUSE'] and temp_start:
                        daily_sec += (ev.timestamp - temp_start)
                        temp_start = None
                
                hours = round(daily_sec / 3600, 2)
                total_hours += hours
                
                # Добавляем строку в таблицу
                ws.cell(row=current_row, column=1).value = tnum if date_idx == 0 else ""
                ws.cell(row=current_row, column=2).value = fio if date_idx == 0 else ""
                ws.cell(row=current_row, column=3).value = pos if date_idx == 0 else ""
                ws.cell(row=current_row, column=4).value = date_str
                ws.cell(row=current_row, column=5).value = f"{s_str}-{e_str}"
                ws.cell(row=current_row, column=6).value = hours
                if include_completed_tasks:
                    ws.cell(row=current_row, column=7).value = tasks_done if date_idx == 0 else ""
                
                # Применяем стили к ячейкам
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = regular_font
                    cell.border = thin_border
                    if col in [6, 7]:
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align

                current_row += 1

            # Итого по сотруднику
            # Колонки F и G должны оставаться отдельными для значений "Часов" и "Выполнено задач".
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell_total_label = ws.cell(row=current_row, column=1)
            cell_total_label.value = f"ИТОГО {fio}:"
            cell_total_label.font = bold_font
            cell_total_label.fill = total_fill
            cell_total_label.alignment = left_align
            cell_total_label.border = thin_border
            
            cell_total_hours = ws.cell(row=current_row, column=6)
            cell_total_hours.value = total_hours
            cell_total_hours.font = bold_font
            cell_total_hours.fill = total_fill
            cell_total_hours.alignment = right_align
            cell_total_hours.border = thin_border
            if include_completed_tasks:
                cell_total_tasks = ws.cell(row=current_row, column=7)
                cell_total_tasks.value = tasks_done
                cell_total_tasks.font = bold_font
                cell_total_tasks.fill = total_fill
                cell_total_tasks.alignment = right_align
                cell_total_tasks.border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Timesheet_{data['start_date']}_{data['end_date']}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        wb.save(response)
        return response
    
    actions = ['go_to_export']
    def go_to_export(self, request, queryset):
        return HttpResponseRedirect("export-report/")
    go_to_export.short_description = "📊 Выгрузить табель"

@admin.register(GlobalSettings)
class SettingsAdmin(admin.ModelAdmin):
    list_display = ('company_name',)